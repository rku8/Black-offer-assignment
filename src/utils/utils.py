import sys
import os
import re
import nltk
import syllapy
import pandas as pd
from src.logger import logging
from openpyxl import load_workbook
from openpyxl.styles import Font
from src.exception import CustomException

def save_file(file_path: str, text: str):
    try:
        with open(file_path, "w") as file:
            file.write(text)
        logging.info(f"Text is successfully saved in {file_path}")
    except Exception as e:
        raise CustomException(e, sys)

def save_excel(data, filename):
    try:
        df = pd.DataFrame(data)
        df.to_excel(filename, index=False, engine='openpyxl')
        wb = load_workbook(filename)
        ws = wb.active
        url_column_index = df.columns.get_loc('URL') + 1

        # Iterate over the rows in the DataFrame to add hyperlinks
        for row in ws.iter_rows(min_row=2, min_col=url_column_index, max_col=url_column_index, max_row=len(df) + 1):
            for cell in row:
                if cell.value:
                    cell.value = f'=HYPERLINK("{cell.value}", "{cell.value}")'
                    cell.font = Font(color='0000FF', underline='single')

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column name

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = (max_length + 2)  # Add some extra space for better readability
            ws.column_dimensions[column_letter].width = adjusted_width
        ws.column_dimensions[ws.cell(row=1, column=url_column_index).column_letter].width = 130  
        wb.save(filename)

    except Exception as e:
        raise CustomException(e, sys)

def read_file(file_path: str):
    try:
        with open(file_path, "r") as file:
            text = file.read()
        return text
    except Exception as e:
        raise CustomException(e, sys)

def stopwords():
    try:
        stopwords_files = os.listdir('data/StopWords')
        stopwords = ""
        for stopwords_file in stopwords_files:
            # stopwords = read_file(r'instructions\StopWords\\' + stopwords_file)
            with open(r'instructions/StopWords//' + stopwords_file, "r", encoding="latin1") as file:
                stopwords += file.read()
        er = stopwords.replace('\n', " ").replace("|", "")
        ertext = re.sub(r'\s+', ' ', er)
        stopwords_list = ertext.lower().split()
        return stopwords_list
    except Exception as e:
        raise CustomException(e, sys)

def positive_words():
    try:
        with open(r'data\MasterDictionary\positive-words.txt', 'r', encoding="latin1") as file:
            positive_words = file.read()
        er = positive_words.replace("\n", " ")
        er = re.sub(r'\s+', ' ', er)
        positive_words_list = er.split()
        return positive_words_list
    except Exception as e:
        raise CustomException(e, sys)

def negative_words():
    try:
        with open(r'data\MasterDictionary\negative-words.txt', 'r', encoding="latin1") as file:
            negative_words = file.read()
        er = negative_words.replace("\n", " ")
        er = re.sub(r'\s+', ' ', er)
        negative_words_list = er.split()
        return negative_words_list
    except Exception as e:
        raise CustomException(e, sys)

def polarity_score(pos_score, neg_score):
    try:
        score = (pos_score - neg_score)/(pos_score + neg_score + 0.000001)
        return score
    except Exception as e:
        raise CustomException(e, sys)

def subjectivity_score(pos_score, neg_score, words_len):
    try:
        score = (pos_score + neg_score)/(words_len + 0.000001)
        return score
    except Exception as e:
        raise CustomException(e, sys)
    
def word_counts(words):
    try:
        nltk_data_path = os.path.join(os.path.dirname(nltk.__file__), 'nltk_data')
        stopwords_path = os.path.join(nltk_data_path, 'corpora', 'stopwords')
        if not os.path.exists(stopwords_path):
            nltk.download('stopwords', download_dir=nltk_data_path)
        from nltk.corpus import stopwords
        stop_words = set(stopwords.words('english'))
        filtered_words = [word for word in words if word.lower() not in stop_words]
        word_count = len(filtered_words)
        return word_count
    except Exception as e:
        raise CustomException(e, sys)

def syllable_count(word):
    try:
        if word.endswith('es') or word.endswith('ed'):
            word = word[:-2] 
        return syllapy.count(word)
    except Exception as e:
        raise CustomException(e, sys)

def syllable_per_words(words):
    try:
         total_words = len(words)
         total_syllables = sum(syllable_count(word) for word in words)
         syllable_per_word = round(total_syllables / total_words, 3)
         return syllable_per_word
    except Exception as e:
        raise CustomException(e, sys)
    
def count_personal_pronouns(text: str):
    try:
         pronouns = ["I", "we", "We", "my", "My", "ours", "Ours", "Us", "us"]
         pattern = r'\b(?:I|we|We|my|My|ours|Ours|us|Us)\b'
         regex = re.compile(pattern, re.IGNORECASE)
         matches = regex.findall(text)
         pronoun_counts =  sum(matches.count(pronoun) for pronoun in pronouns)
         return pronoun_counts
    
    except Exception as e:
        raise CustomException(e, sys)
    
def average_word_length(words):
    try:
        # Initialize counters
        total_characters = 0
        total_words = 0
        
        # Iterate over words to calculate total characters and word count
        for word in words:
            if word.isalpha():
                total_characters += len(word)
                total_words += 1
        
        # Calculate the average word length
        average_length = total_characters / total_words if total_words > 0 else 0
        
        return average_length
    except Exception as e:
        raise CustomException(e, sys)

def get_input_val(df, filepath):
    try:
        _, filename = os.path.split(filepath)
        url_id = filename.split('.txt')[0]
        url = df[df['URL_ID'] == url_id]['URL'].values[0]
        return url_id, url
    
    except Exception as e:
        raise CustomException(e, sys)