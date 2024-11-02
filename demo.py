import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Example list of dictionaries
data = [
    {
        "URL_ID": 1,
        "URL": "https://www.example.com",
        "POSITIVE SCORE": 10,
        "NEGATIVE SCORE": 2,
        "POLARITY SCORE": 0.8,
        "SUBJECTIVITY SCORE": 0.5,
        "AVG SENTENCE LENGTH": 15,
        "PERCENTAGE OF COMPLEX WORDS": 0.3,
        "FOG INDEX": 12,
        "AVG NUMBER OF WORDS PER SENTENCE": 20,
        "COMPLEX WORD COUNT": 5,
        "WORD COUNT": 200,
        "SYLLABLE PER WORD": 1.5,
        "PERSONAL PRONOUNS": 3,
        "AVG WORD LENGTH": 4.5
    },
    {
        "URL_ID": 2,
        "URL": "https://www.google.com",
        "POSITIVE SCORE": 8,
        "NEGATIVE SCORE": 1,
        "POLARITY SCORE": 0.9,
        "SUBJECTIVITY SCORE": 0.4,
        "AVG SENTENCE LENGTH": 12,
        "PERCENTAGE OF COMPLEX WORDS": 0.2,
        "FOG INDEX": 10,
        "AVG NUMBER OF WORDS PER SENTENCE": 18,
        "COMPLEX WORD COUNT": 4,
        "WORD COUNT": 180,
        "SYLLABLE PER WORD": 1.4,
        "PERSONAL PRONOUNS": 2,
        "AVG WORD LENGTH": 4.3
    }
]

# Convert the list of dictionaries to a pandas DataFrame
df = pd.DataFrame(data)

# Specify the file name
filename = 'output_with_clickable.xlsx'

# Write the DataFrame to an Excel file without styling first
df.to_excel(filename, index=False, engine='openpyxl')

# Load the workbook and select the active worksheet
wb = load_workbook(filename)
ws = wb.active

# Find the index of the 'URL' column
url_column_index = df.columns.get_loc('URL') + 1

# Iterate over the rows in the DataFrame to add hyperlinks
for row in ws.iter_rows(min_row=2, min_col=url_column_index, max_col=url_column_index, max_row=len(df) + 1):
    for cell in row:
        if cell.value:
            cell.value = f'=HYPERLINK("{cell.value}", "{cell.value}")'
            cell.font = Font(color='0000FF', underline='single')

# Adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter  # Get the column name

    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 3)
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
wb.save(filename)

print(f"Data has been written to {filename} with URLs as clickable hyperlinks and adjusted column widths.")
